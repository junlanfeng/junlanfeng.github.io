#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/build_pub_json.py - 从 pub.xlsx 提取论文数据，输出 publications.json
并将数据内联注入到 publications.html（支持 file:// 协议直接打开）
"""

import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# 项目根目录
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(PROJECT_ROOT, 'data', 'pub.xlsx')
OUTPUT_PATH = os.path.join(PROJECT_ROOT, 'data', 'publications.json')
HTML_PATH = os.path.join(PROJECT_ROOT, 'publications.html')

# 列名关键词映射（用关键词匹配，避免列名细微差异）
COLUMN_MAP = {
    'type': ['类型（会议/期刊/专著/专利等）', '类型'],
    'title_zh': ['中文题目', '中文标题', '题（中文）'],
    'title_en': ['英文题目', '英文标题', '题（英文）'],
    'source_zh': ['英文发表来源', '英文来源'],
    'source_en': ['英文发表来源', '英文来源'],
    'year': ['时间（年份）', '时间', '年份'],
    'citations': ['引用次数', '被引'],
    'impact_factor': ['影响因子（学术期刊）', '影响因子'],
    'bibtex': ['BibTeX', 'bibtex', 'BIBTEX'],
    'pdf': ['pdf原文', 'PDF', 'pdf', '原文'],
}


def find_column(headers, keywords):
    """根据关键词模糊匹配列名"""
    for h in headers:
        if h is None:
            continue
        h_str = str(h).strip()
        for kw in keywords:
            if kw in h_str:
                return h_str
    return None


def extract_year(val):
    """从年份单元格提取4位年份。
    兼容: 2020(数字) / 2020.0(Excel数字) / '2020年' / '2020-05' / datetime字符串 等。
    解析失败返回 0（排序时落到最后）。
    """
    if val is None:
        return 0
    m = re.search(r'(19|20)\d{2}', str(val))
    return int(m.group(0)) if m else 0


def validate_url(val):
    """校验是否为有效 URL（http/https 开头），无效则返回空字符串。"""
    if not val:
        return ''
    s = str(val).strip()
    if re.match(r'^https?://', s, re.IGNORECASE):
        return s
    return ''


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"错误: 找不到 {XLSX_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"表头共 {len(headers)} 列:")
    for i, h in enumerate(headers):
        print(f"  Col {i}: {h}")

    # 匹配列名
    col_indices = {}
    for key, keywords in COLUMN_MAP.items():
        col_name = find_column(headers, keywords)
        if col_name:
            col_indices[key] = headers.index(col_name)
            print(f"  匹配 {key} -> 列 '{col_name}' (index {col_indices[key]})")
        else:
            print(f"  ⚠️ 未找到 {key} 列")
            col_indices[key] = None

    type_col = col_indices.get('type')
    if type_col is None:
        print("❌ 未找到类型列，无法筛选论文")
        sys.exit(1)

    # 提取论文数据
    papers = []
    total_rows = 0
    matched = 0
    skipped_reasons = {}  # 统计跳过原因

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        total_rows += 1
        type_val = row[type_col] if type_col < len(row) else None

        if type_val is None:
            reason = 'type为None'
            skipped_reasons[reason] = skipped_reasons.get(reason, 0) + 1
            print(f"  [行{row_idx}] 跳过: type为None, title_zh={str(row[col_indices.get('title_zh', -1)] or '')[:30] if col_indices.get('title_zh') is not None and col_indices['title_zh'] < len(row) else '?'}")
            continue

        type_str = str(type_val).strip()

        if type_str != '论文':
            # 检查是否包含"论文"但有多余字符
            if '论文' in type_str:
                reason = f'含"论文"但不完全匹配: repr={repr(type_val)}'
                print(f"  [行{row_idx}] ⚠️ 含论文但不匹配: repr={repr(type_val)}, 长度={len(type_str)}, hex={type_str.encode('utf-8').hex()}")
            else:
                reason = f'type={type_str}'
            skipped_reasons[reason] = skipped_reasons.get(reason, 0) + 1
            continue

        matched += 1

        def get_val(key):
            idx = col_indices.get(key)
            if idx is None or idx >= len(row):
                return ''
            val = row[idx]
            return '' if val is None else str(val).strip()

        papers.append({
            'title_zh': get_val('title_zh'),
            'title_en': get_val('title_en'),
            'source_zh': get_val('source_zh'),
            'source_en': get_val('source_en'),
            'year': get_val('year'),
            'citations': get_val('citations'),
            'impact_factor': get_val('impact_factor'),
            'bibtex': get_val('bibtex'),
            'pdf': validate_url(get_val('pdf')),
        })

    wb.close()

    # 统计汇总
    print(f"\n{'='*50}")
    print(f"总数据行: {total_rows}")
    print(f"匹配论文: {matched}")
    print(f"跳过: {total_rows - matched}")
    print(f"\n跳过原因统计:")
    for reason, count in sorted(skipped_reasons.items(), key=lambda x: -x[1]):
        print(f"  [{count}次] {reason}")
    print(f"{'='*50}")

    # 按年份倒序排列（JSON 文件和注入 HTML 共用此列表，两边都是倒序）
    # Python sort 是稳定排序：年份相同的记录保持 Excel 原始顺序；年份缺失(0)排最后
    papers.sort(key=lambda x: extract_year(x.get('year')), reverse=True)

    # 验证排序键是否真正解析出年份（防止全部回退为 0）
    parsed_years = [extract_year(p['year']) for p in papers]
    valid_years = [y for y in parsed_years if y]
    print(f"📅 年份倒序: 最新 {max(valid_years) if valid_years else '?'} ~ 最早 {min(valid_years) if valid_years else '?'}"
          f"，成功解析 {len(valid_years)}/{len(papers)} 条")
    if len(valid_years) < len(papers):
        print(f"⚠️ {len(papers) - len(valid_years)} 条未解析到年份（已排到列表末尾），前3条:")
        for p in papers[-3:]:
            print(f"   year={repr(p['year'])}, title_zh={p['title_zh'][:40]}")
    print(f"   排序后前3条年份: {[extract_year(p['year']) for p in papers[:3]]}")

    # 输出 JSON
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(papers, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 导出 {len(papers)} 篇论文 -> {OUTPUT_PATH}")

    # 打印前3条预览
    for i, p in enumerate(papers[:3]):
        print(f"\n--- 预览 {i+1} ---")
        print(f"  title_zh: {p['title_zh'][:50]}")
        print(f"  title_en: {p['title_en'][:50]}")
        print(f"  source_zh: {p['source_zh']}")
        print(f"  source_en: {p['source_en']}")
        print(f"  year: {p['year']}")
        print(f"  citations: {p['citations']}")
        print(f"  impact_factor: {p['impact_factor']}")
        print(f"  bibtex: {p['bibtex'][:50] if p['bibtex'] else '(空)'}...")
        print(f"  pdf: {p['pdf'][:50] if p['pdf'] else '(空)'}...")

    # 将数据内联注入到 publications.html
    inject_into_html(papers)


def inject_into_html(papers):
    """将论文数据内联注入到 publications.html，支持 file:// 协议直接打开"""
    if not os.path.exists(HTML_PATH):
        print(f"⚠️ publications.html 不存在，跳过注入")
        return

    with open(HTML_PATH, 'r', encoding='utf-8') as f:
        html = f.read()

    # 生成内联 JS
    json_str = json.dumps(papers, ensure_ascii=False, separators=(',', ':'))
    inline_js = f'var publicationsData = {json_str};'
    inline_script = f'<script>\n{inline_js}\n</script>\n    '

    # 删除已有的 publicationsData 定义
    patterns = [
        r'<script>\s*var\s+publicationsData\s*=\s*[^;]+;\s*</script>',
        r'<script>\s*const\s+publicationsData\s*=\s*[^;]+;\s*</script>',
        r'<script>\s*let\s+publicationsData\s*=\s*[^;]+;\s*</script>',
        r'var\s+publicationsData\s*=\s*[^;]+;',
        r'const\s+publicationsData\s*=\s*[^;]+;',
        r'let\s+publicationsData\s*=\s*[^;]+;',
    ]
    for pattern in patterns:
        html = re.sub(pattern, '', html, flags=re.DOTALL)

    # 清理多余空行
    html = re.sub(r'\n\s*\n\s*\n', '\n\n', html)

    # 在 pub-render.js 之前插入
    script_tag = '<script src="js/pub-render.js">'
    if script_tag in html:
        html = html.replace(script_tag, inline_script + script_tag)
        print(f"✅ 内联数据已注入到 publications.html")
    else:
        print(f"⚠️ 未找到 pub-render.js 引用，未注入内联数据")

    with open(HTML_PATH, 'w', encoding='utf-8') as f:
        f.write(html)

    size_kb = os.path.getsize(HTML_PATH) / 1024
    print(f"📄 publications.html 大小: {size_kb:.1f} KB")


if __name__ == '__main__':
    main()
